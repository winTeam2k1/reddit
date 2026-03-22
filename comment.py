from __future__ import annotations

import json
import re
import tempfile
import time
import urllib.parse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from post import (
    CHROME_USER_AGENT,
    VIEWPORT,
    RedditScreenshotError,
    browser_can_show_ui,
    capture_post_slices,
    dismiss_common_popups,
    ensure_not_blocked,
    get_post_bounds,
    hide_fixed_overlays,
    import_playwright,
    launch_browser,
    normalize_reddit_url,
    sanitize_output_name,
    stitch_pngs,
)

OUTPUT_DIR = Path("comments")
EXCEL_PATH = Path("data.xlsx")
CONFIG_PATH = Path("commentConfig.json")
DELAY_BETWEEN_COMMENTS_SECONDS = 3.0
ERROR_REPORT_PATH = OUTPUT_DIR / "comment_errors.txt"


@dataclass(frozen=True)
class CommentJob:
    row_number: int
    output_name: str
    link: str


@dataclass(frozen=True)
class CommentConfig:
    start: int
    end: int | None


@dataclass(frozen=True)
class CommentFailure:
    row_number: int
    output_name: str
    link: str
    message: str


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def select_comment_sheet(workbook):
    if len(workbook.worksheets) >= 2:
        return workbook.worksheets[1]
    if "Sheet2" in workbook.sheetnames:
        return workbook["Sheet2"]
    raise RedditScreenshotError("Khong tim thay sheet 2 trong data.xlsx.")


def load_jobs_from_excel(path: Path) -> list[CommentJob]:
    if not path.exists():
        raise RedditScreenshotError(f"Khong tim thay file Excel: {path.resolve()}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = select_comment_sheet(workbook)
        jobs: list[CommentJob] = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            output_name = cell_to_text(row[0]) if len(row) >= 1 else ""
            link = cell_to_text(row[4]) if len(row) >= 5 else ""

            if not any(cell_to_text(value) for value in row):
                continue
            if not output_name or not link:
                print(f"Bo qua dong {row_number}: thieu cot A hoac cot E.")
                continue

            jobs.append(
                CommentJob(
                    row_number=row_number,
                    output_name=output_name,
                    link=normalize_reddit_url(link),
                )
            )
    finally:
        workbook.close()

    if not jobs:
        raise RedditScreenshotError("Khong co dong comment hop le nao trong sheet 2.")
    return jobs


def parse_config_int(config: dict[str, object], key: str) -> int | None:
    value = config.get(key)
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, int):
        raise RedditScreenshotError(f"Gia tri '{key}' trong {CONFIG_PATH} phai la so nguyen.")
    if value < 1:
        raise RedditScreenshotError(f"Gia tri '{key}' trong {CONFIG_PATH} phai >= 1.")
    return value


def load_comment_config(path: Path) -> CommentConfig:
    if not path.exists():
        raise RedditScreenshotError(f"Khong tim thay file config: {path.resolve()}")

    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RedditScreenshotError(f"File {path} khong phai JSON hop le: {exc}") from exc

    if not isinstance(config, dict):
        raise RedditScreenshotError(f"File {path} phai la mot JSON object.")

    start = parse_config_int(config, "start")
    end = parse_config_int(config, "end")
    if start is None:
        start = 1
    if end is not None and end < start:
        raise RedditScreenshotError("Gia tri 'end' phai lon hon hoac bang 'start'.")

    return CommentConfig(start=start, end=end)


def filter_jobs_by_config(jobs: list[CommentJob], config: CommentConfig) -> list[CommentJob]:
    first_row = jobs[0].row_number
    last_row = jobs[-1].row_number
    effective_end = last_row if config.end is None else min(config.end, last_row)

    filtered = [
        job for job in jobs if config.start <= job.row_number <= effective_end
    ]
    if not filtered:
        raise RedditScreenshotError(
            f"Khong co dong comment hop le nao trong khoang Excel row {config.start} den {effective_end}."
        )

    print(
        f"Chay comment tu dong Excel {config.start} den {effective_end} "
        f"(du lieu hop le tu {first_row} den {last_row})."
    )
    return filtered


def extract_comment_target(url: str) -> tuple[str | None, str | None]:
    path = urllib.parse.urlparse(url).path
    match = re.match(r"^(/r/[^/]+/comments/[^/]+/comment/([^/]+)/?)/?$", path)
    if match:
        permalink = match.group(1)
        if not permalink.endswith("/"):
            permalink += "/"
        return permalink, match.group(2)
    return None, None


def make_output_path(name: str) -> Path:
    stem = sanitize_output_name(name)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    candidate = OUTPUT_DIR / f"#{stem}.png"
    index = 2
    while candidate.exists():
        candidate = OUTPUT_DIR / f"#{stem}_{index}.png"
        index += 1
    return candidate


def find_target_comment(page, permalink: str | None, comment_id: str | None):
    target = {"permalink": permalink, "commentId": comment_id}
    handle = page.evaluate_handle(
        """
        (target) => {
          const comments = Array.from(document.querySelectorAll("shreddit-comment"));
          return comments.find((node) => {
            const permalink = node.getAttribute("permalink") || "";
            const thingId = node.getAttribute("thingid") || "";
            if (target.permalink && permalink === target.permalink) {
              return true;
            }
            if (target.commentId) {
              return permalink.includes(`/comment/${target.commentId}/`) ||
                thingId === `t1_${target.commentId}`;
            }
            return false;
          }) || null;
        }
        """,
        target,
    )
    return handle.as_element()


def wait_for_target_comment(page, permalink: str | None, comment_id: str | None):
    for _ in range(30):
        element = find_target_comment(page, permalink, comment_id)
        if element is not None:
            return element
        page.wait_for_timeout(500)
    raise RedditScreenshotError("Khong tim thay comment duoc highlight trong trang.")


def take_comment_screenshot_once(url: str, output_path: Path, *, headless: bool) -> Path:
    sync_playwright, PlaywrightTimeoutError, PlaywrightError = import_playwright()

    with sync_playwright() as playwright:
        browser = launch_browser(playwright, headless=headless)
        context = browser.new_context(
            viewport=VIEWPORT,
            device_scale_factor=2,
            color_scheme="light",
            locale="en-US",
            user_agent=CHROME_USER_AGENT,
            extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
        )
        page = context.new_page()
        page.set_default_timeout(45_000)
        page.add_init_script(
            """
            Object.defineProperty(navigator, 'webdriver', {
              get: () => undefined
            });
            """
        )
        try:
            page.goto(url, wait_until="domcontentloaded")
            try:
                page.wait_for_url(re.compile(r".*/comments/.*"), timeout=12_000)
            except PlaywrightTimeoutError:
                pass
            try:
                page.wait_for_load_state("networkidle", timeout=8_000)
            except PlaywrightTimeoutError:
                pass

            ensure_not_blocked(page)
            dismiss_common_popups(page)

            resolved_url = normalize_reddit_url(page.url)
            permalink, comment_id = extract_comment_target(resolved_url)
            if not comment_id:
                permalink, comment_id = extract_comment_target(url)
            if not comment_id:
                raise RedditScreenshotError(
                    "Khong phan tich duoc comment id tu link nay. Hay dung link comment/permalink."
                )

            comment = wait_for_target_comment(page, permalink, comment_id)
            comment.scroll_into_view_if_needed(timeout=5_000)
            page.wait_for_timeout(500)

            hide_fixed_overlays(page, comment)
            page.wait_for_timeout(300)
            bounds = get_post_bounds(comment)

            with tempfile.TemporaryDirectory(prefix="reddit_comment_") as temp_dir_name:
                temp_dir = Path(temp_dir_name)
                parts = capture_post_slices(page, bounds, temp_dir)
                if len(parts) == 1:
                    output_path.write_bytes(parts[0].read_bytes())
                else:
                    stitch_pngs(parts, output_path)

            return output_path
        except PlaywrightTimeoutError as exc:
            raise RedditScreenshotError(
                "Het thoi gian cho khi tai Reddit hoac tim comment duoc highlight."
            ) from exc
        except PlaywrightError as exc:
            raise RedditScreenshotError(f"Playwright loi khi chup anh comment: {exc}") from exc
        finally:
            context.close()
            browser.close()


def take_comment_screenshot(url: str, output_path: Path) -> Path:
    attempts = [True]
    if browser_can_show_ui():
        attempts.append(False)

    last_error: RedditScreenshotError | None = None
    for headless in attempts:
        try:
            return take_comment_screenshot_once(url, output_path, headless=headless)
        except RedditScreenshotError as exc:
            last_error = exc
            if "Reddit dang chan request" not in str(exc) or not headless:
                raise

    if last_error is None:
        raise RedditScreenshotError("Khong the chup comment Reddit.")
    raise RedditScreenshotError(
        f"{last_error} Thu thu lai tren mang khac, dang nhap Reddit, hoac chay browser hien hinh."
    )


def write_failure_report(failures: list[CommentFailure]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    lines = ["Dòng\tSTT\t\tLink"]
    for index, failure in enumerate(failures, start=1):
        lines.append(f"{index}\t\t{failure.output_name}\t\t{failure.link}")

    ERROR_REPORT_PATH.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return ERROR_REPORT_PATH


def report_failures(failures: list[CommentFailure]) -> None:
    if not failures:
        return

    report_path = write_failure_report(failures)
    print(f"Da ghi log loi tai: {report_path.resolve()}")


def capture_jobs(jobs: list[CommentJob]) -> list[CommentFailure]:
    failures: list[CommentFailure] = []
    total = len(jobs)

    for index, job in enumerate(jobs, start=1):
        try:
            output_path = make_output_path(job.output_name)
            final_path = take_comment_screenshot(job.link, output_path)
            print(f"[{index}/{total}] Dong Excel {job.row_number} | Ten file: #{job.output_name}")
            print(f"Da luu anh comment tai: {final_path.resolve()}")
        except RedditScreenshotError as exc:
            print(f"[{index}/{total}] Dong Excel {job.row_number} - loi: {exc}")
            failures.append(
                CommentFailure(
                    row_number=job.row_number,
                    output_name=job.output_name,
                    link=job.link,
                    message=str(exc),
                )
            )

        if index < total and DELAY_BETWEEN_COMMENTS_SECONDS > 0:
            time.sleep(DELAY_BETWEEN_COMMENTS_SECONDS)

    return failures


def main() -> int:
    try:
        jobs = load_jobs_from_excel(EXCEL_PATH)
        config = load_comment_config(CONFIG_PATH)
        jobs = filter_jobs_by_config(jobs, config)
        print(f"Tim thay {len(jobs)} link comment hop le trong {EXCEL_PATH.resolve()}")
        failures = capture_jobs(jobs)
        if failures:
            report_failures(failures)
            print(f"Hoan tat voi {len(failures)} dong loi.")
            return 1
        print("Hoan tat tat ca comment.")
        return 0
    except KeyboardInterrupt:
        print("\nDa huy.")
        return 130
    except RedditScreenshotError as exc:
        print(f"Loi: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
